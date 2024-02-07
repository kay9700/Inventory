import openpyxl
from datetime import date

#Open workbook
workbook = openpyxl.load_workbook('example.xlsx')

#Open specific sheet
worksheet = workbook['Sheet1']

def main():
    update_header()

def update_header():
    firstRowArr = [];
    today = date.today()
    d1 = today.strftime("%d/%m/%Y")
    #print("d1 =", d1)
    #print("Today's date:", today)

    for col in worksheet.iter_cols():
        for cell in col:
            if ((cell.value == '') or (cell.value == d1)): 
                if cell.value == '' : 
                    worksheet.cell(value=d1)
                    workbook.save('example.xlsx')
                    print("poner fecha")
            #firstRowArr.append(cell.value)

    """ for i in range(5): 
        value = worksheet.cell(row=1, column=i).value
        firstRowArr.append(value) """

    print(firstRowArr)

    #value = worksheet.cell(row=1, column=1).value
    worksheet.cell(row=1, column=2, value=d1)
    workbook.save('example.xlsx')
    

main()

# Read a single cell
#value = worksheet.cell(row=1, column=1).value
#print(value)


# Read a range of cells

for row in worksheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3):
    for cell in row:
        print("cell.value")

    

